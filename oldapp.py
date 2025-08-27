import os
from datetime import datetime
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import ttkbootstrap as tb
import platform
import subprocess
import requests
import threading
import webbrowser
import json
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
import psutil
import time # Importar time para usar sleep

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

# --- Configurações Globais (podem ser movidas para um arquivo de config.py) ---
TIPOS = ["VT", "Exp", "Meia", "Meia V", "QR Code"]
VERSAO_LOCAL = "1.0.0"

def brl(v):
    s = f"{float(v):,.2f}"
    s = s.replace(",", "X").replace(".", ",").replace("X", ".")
    return f"R$ {s}"

def tarifa(linha_str, tipo):
    linha = str(linha_str).strip()
    if linha == "970":
        return 7.40 if tipo in ("VT", "Exp", "QR Code") else 3.70
    return 4.60 if tipo in ("VT", "Exp", "QR Code") else 2.30

def validar_data(data_ddmmaa):
    try:
        datetime.strptime(data_ddmmaa, "%d/%m/%Y")
        return True
    except ValueError:
        return False

class ControleBilhetagemApp:
    def __init__(self, master):
        self.master = master
        master.title("Controle de Bilhetagem")
        # Removendo geometry fixo para permitir redimensionamento
        # master.geometry("1280x700")

        self._create_widgets()
        self._bind_events()

    def _create_widgets(self):
        # Frame Top
        frame_top = ttk.Frame(self.master, padding=10)
        frame_top.pack(fill="x")

        ttk.Label(frame_top, text="Data (DD/MM/AAAA):").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        self.entry_data = ttk.Entry(frame_top, width=12)
        self.entry_data.insert(0, datetime.now().strftime("%d/%m/%Y"))
        self.entry_data.grid(row=0, column=1, padx=5, pady=5)

        ttk.Label(frame_top, text="Linha:").grid(row=0, column=2, padx=5, pady=5, sticky="w")
        self.entry_linha = ttk.Entry(frame_top, width=10)
        self.entry_linha.grid(row=0, column=3, padx=5, pady=5)

        ttk.Label(frame_top, text="Prefixo:").grid(row=0, column=4, padx=5, pady=5, sticky="w")
        self.entry_prefixo = ttk.Entry(frame_top, width=10)
        self.entry_prefixo.grid(row=0, column=5, padx=5, pady=5)

        ttk.Label(frame_top, text="Turno:").grid(row=0, column=6, padx=5, pady=5, sticky="w")
        self.turno_var = tk.StringVar()
        self.combo_turno = ttk.Combobox(frame_top, textvariable=self.turno_var, values=["1º Turno", "2º Turno", "3º Turno"], state="readonly", width=12)
        self.combo_turno.current(0)
        self.combo_turno.grid(row=0, column=7, padx=5, pady=5)

        ttk.Label(frame_top, text="Empresa:").grid(row=1, column=0, padx=5, pady=5, sticky="w")
        self.empresa_var = tk.StringVar()
        self.combo_empresa = ttk.Combobox(frame_top, textvariable=self.empresa_var, values=["MONTE CRISTO", "TRANSPORTES CANADÁ"], state="readonly", width=20)
        self.combo_empresa.current(0)
        self.combo_empresa.grid(row=1, column=1, padx=5, pady=5)

        # Frame Quantidade
        frame_qtd = ttk.Frame(self.master, padding=10)
        frame_qtd.pack(fill="x")

        self.entries_tipo = {}
        col = 0
        for tipo in TIPOS:
            ttk.Label(frame_qtd, text=tipo + ":").grid(row=0, column=col, padx=5, pady=5, sticky="e")
            e = ttk.Entry(frame_qtd, width=6)
            e.grid(row=0, column=col+1, padx=5, pady=5, sticky="w")
            e.insert(0, "0")
            self.entries_tipo[tipo] = e
            col += 2

        # Tabela
        colunas = ["Data", "Linha", "Prefixo", "Turno"] + TIPOS + ["Empresa"]
        self.tabela = ttk.Treeview(self.master, columns=colunas, show="headings", height=16)
        for c in colunas:
            self.tabela.heading(c, text=c)
            self.tabela.column(c, width=110, anchor="center")
        self.tabela.pack(fill="both", expand=True, padx=10, pady=10)

        # Frame Botões (parte inferior)
        frame_btn_bottom = ttk.Frame(self.master, padding=10)
        frame_btn_bottom.pack()

        tb.Button(frame_btn_bottom, text="Salvar Trabalho", bootstyle="success", command=self.salvar_trabalho).grid(row=0, column=0, padx=8)
        tb.Button(frame_btn_bottom, text="Carregar Trabalho", bootstyle="info", command=self.carregar_trabalho).grid(row=0, column=1, padx=8)
        tb.Button(frame_btn_bottom, text="Adicionar", bootstyle="success", command=self.adicionar).grid(row=0, column=2, padx=8)
        tb.Button(frame_btn_bottom, text="Remover Selecionado", bootstyle="danger", command=self.remover_selecionado).grid(row=0, column=3, padx=8)
        tb.Button(frame_btn_bottom, text="Exportar PDF", bootstyle="info", command=self.exportar_pdf).grid(row=0, column=4, padx=8)
        tb.Button(frame_btn_bottom, text="Exportar Excel", bootstyle="success", command=self.exportar_excel).grid(row=0, column=5, padx=8)
        ttk.Button(frame_btn_bottom, text="Verificar Atualizações", command=self.verificar_atualizacao).grid(row=0, column=6, padx=8)
        ttk.Label(self.master, text="© 2025 giovanne_martins", foreground="gray").pack(side="bottom", pady=5)

    def _bind_events(self):
        widgets_ordem = [self.entry_data, self.entry_linha, self.entry_prefixo] + list(self.entries_tipo.values()) + [self.combo_turno, self.combo_empresa]
        for w in widgets_ordem:
            w.bind("<Return>", self._navegar_enter)

    def _navegar_enter(self, event):
        widget = event.widget
        widgets_ordem = [self.entry_data, self.entry_linha, self.entry_prefixo] + list(self.entries_tipo.values()) + [self.combo_turno, self.combo_empresa]
        try:
            idx = widgets_ordem.index(widget)
            if idx < len(widgets_ordem) - 1:
                widgets_ordem[idx + 1].focus_set()
            else:
                self.adicionar()
        except ValueError:
            pass

    def verificar_atualizacao(self):
        url_versao = "https://raw.githubusercontent.com/timontecristo-design/mcbilhetagem/main/update.txt"
        try:
            r = requests.get(url_versao)
            r.raise_for_status()
            versao_nova = r.text.strip()
            if versao_nova != VERSAO_LOCAL:
                messagebox.showinfo("Atualização", f"Nova versão disponível: {versao_nova}")
                if messagebox.askyesno("Confirmar Atualização", "Deseja baixar e instalar a nova versão agora?"):
                    self._baixar_e_instalar(versao_nova)
            else:
                messagebox.showinfo("Sem Atualizações", "Você está com a versão mais recente.")
        except requests.exceptions.RequestException as e:
            messagebox.showerror("Erro", f"Erro ao verificar atualização: {str(e)}")

    def _baixar_e_instalar(self, versao):
        url_instalador = "https://raw.githubusercontent.com/timontecristo-design/mcbilhetagem/main/mysetup.exe"
        desktop = os.path.join(os.path.expanduser("~"), "Desktop")
        nome_arquivo = os.path.join(desktop, f"instalador_v{versao}.exe")
        
        try:
            progresso_win = tk.Toplevel()
            progresso_win.title("Baixando Atualização")
            progresso_win.geometry("400x100")
            ttk.Label(progresso_win, text=f"Baixando instalador v{versao}...").pack(pady=10)
            barra = ttk.Progressbar(progresso_win, orient="horizontal", length=350, mode="determinate")
            barra.pack(pady=10)

            def download():
                try:
                    with requests.get(url_instalador, stream=True, timeout=10) as r:
                        if r.status_code == 404:
                            progresso_win.destroy()
                            messagebox.showerror("Erro", "Instalador não encontrado no servidor.")
                            return
                        r.raise_for_status()
                        total = int(r.headers.get("content-length", 0))
                        downloaded = 0
                        chunk_size = 8192
                        with open(nome_arquivo, "wb") as f:
                            for chunk in r.iter_content(chunk_size=chunk_size):
                                if chunk:
                                    f.write(chunk)
                                    downloaded += len(chunk)
                                    barra["value"] = downloaded / total * 100 if total else 100
                                    progresso_win.update_idletasks()
                    # Feedback visual: manter a janela de progresso por um breve período
                    ttk.Label(progresso_win, text="Download Concluído!").pack(pady=5)
                    progresso_win.update_idletasks()
                    time.sleep(1) # Espera 1 segundo
                    progresso_win.destroy()
                    messagebox.showinfo("Instalador Baixado", f"Instalador v{versao} baixado com sucesso!")
                    if platform.system() == "Windows":
                        subprocess.Popen([nome_arquivo], shell=True) # Use nome_arquivo aqui
                    elif platform.system() == "Darwin":
                        subprocess.Popen(["open", nome_arquivo])
                    else:
                        subprocess.Popen(["xdg-open", nome_arquivo])
                except requests.exceptions.Timeout:
                    progresso_win.destroy()
                    messagebox.showerror("Erro", "Tempo de conexão esgotado durante o download.")
                except requests.exceptions.ConnectionError:
                    progresso_win.destroy()
                    messagebox.showerror("Erro", "Não foi possível conectar ao servidor do instalador.")
                except Exception as e:
                    progresso_win.destroy()
                    messagebox.showerror("Erro", f"Falha ao baixar o instalador: {str(e)}")

            threading.Thread(target=download, daemon=True).start()

        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao iniciar download: {str(e)}")

    def salvar_trabalho(self):
        data = self.entry_data.get().strip()
        linha = self.entry_linha.get().strip()
        prefixo = self.entry_prefixo.get().strip()
        turno = self.turno_var.get().strip()
        empresa = self.empresa_var.get().strip()

        valores = {}
        for t in TIPOS:
            try:
                valores[t] = int(self.entries_tipo[t].get())
            except ValueError:
                valores[t] = 0

        dados = {
            "data": data,
            "linha": linha,
            "prefixo": prefixo,
            "turno": turno,
            "empresa": empresa,
            "valores": valores
        }

        pasta_destino = filedialog.askdirectory(title="Escolha o local para salvar o trabalho")

        if not pasta_destino:
            messagebox.showwarning("Aviso", "Nenhum diretório foi selecionado. Trabalho não salvo.")
            return

        nome_arquivo = os.path.join(pasta_destino, "trabalho_salvo.json")

        try:
            with open(nome_arquivo, 'w') as f:
                json.dump(dados, f, indent=4)
            messagebox.showinfo("Sucesso", f"Trabalho salvo com sucesso em {nome_arquivo}")
        except IOError as e:
            messagebox.showerror("Erro de E/S", f"Erro ao salvar o trabalho: {str(e)}")
        except Exception as e:
            messagebox.showerror("Erro", f"Erro inesperado ao salvar o trabalho: {str(e)}")

    def carregar_trabalho(self):
        arquivo_destino = filedialog.askopenfilename(title="Escolha o arquivo para carregar o trabalho",
                                                     filetypes=[("JSON Files", "*.json")])

        if not arquivo_destino:
            messagebox.showwarning("Aviso", "Nenhum arquivo foi selecionado. Trabalho não carregado.")
            return

        try:
            with open(arquivo_destino, 'r') as f:
                dados = json.load(f)

            self.entry_data.delete(0, "end")
            self.entry_data.insert(0, dados["data"])
            self.entry_linha.delete(0, "end")
            self.entry_linha.insert(0, dados["linha"])
            self.entry_prefixo.delete(0, "end")
            self.entry_prefixo.insert(0, dados["prefixo"])
            self.turno_var.set(dados["turno"])
            self.empresa_var.set(dados["empresa"])

            for t in TIPOS:
                self.entries_tipo[t].delete(0, "end")
                self.entries_tipo[t].insert(0, str(dados["valores"].get(t, 0)))

            messagebox.showinfo("Sucesso", "Trabalho carregado com sucesso!")

        except FileNotFoundError:
            messagebox.showerror("Erro", "Arquivo não encontrado.")
        except json.JSONDecodeError:
            messagebox.showerror("Erro", "Formato de arquivo JSON inválido.")
        except KeyError as e:
            messagebox.showerror("Erro", f"Dados ausentes no arquivo JSON: {e}")
        except Exception as e:
            messagebox.showerror("Erro", f"Erro inesperado ao carregar o trabalho: {str(e)}")

    def _limpar_campos(self):
        self.entry_linha.delete(0, "end")
        self.entry_prefixo.delete(0, "end")
        for t in TIPOS:
            self.entries_tipo[t].delete(0, "end")
            self.entries_tipo[t].insert(0, "0")
        self.combo_turno.current(0)
        self.combo_empresa.current(0)
        self.entry_prefixo.focus_set()

    def adicionar(self):
        data = self.entry_data.get().strip()
        linha = self.entry_linha.get().strip()
        prefixo = self.entry_prefixo.get().strip()
        turno = self.turno_var.get().strip()
        empresa = self.empresa_var.get().strip()

        if not data or not linha or not prefixo:
            messagebox.showwarning("Atenção", "Preencha Data, Linha e Prefixo.")
            return
        if not validar_data(data):
            messagebox.showerror("Erro", "Data inválida. Use o formato DD/MM/AAAA.")
            return

        valores = {}
        for t in TIPOS:
            try:
                valor_str = self.entries_tipo[t].get().strip()
                if not valor_str: # Treat empty string as 0
                    valores[t] = 0
                else:
                    valores[t] = int(valor_str)
            except ValueError:
                messagebox.showerror("Erro de Entrada", f"Valor inválido para {t}. Por favor, insira um número inteiro.")
                self.entries_tipo[t].focus_set()
                return # Stop execution if invalid input

        for item in self.tabela.get_children():
            v = self.tabela.item(item, "values")
            if v[1] == linha and v[2] == prefixo and v[3] == turno:
                for t in TIPOS:
                    # Ensure existing values are also treated as integers
                    try:
                        existing_val = int(str(v[4 + TIPOS.index(t)]).split(' ')[0]) # Extract number before '(' if present
                    except ValueError:
                        existing_val = 0 # Default to 0 if existing value is not a valid number
                    valores[t] += existing_val
                self.tabela.item(item, values=(data, linha, prefixo, turno,
                                          valores["VT"], valores["Exp"], valores["Meia"], valores["Meia V"], valores["QR Code"],
                                          empresa))
                self._limpar_campos()
                return

        self.tabela.insert("", "end", values=(data, linha, prefixo, turno,
                                         valores["VT"], valores["Exp"], valores["Meia"], valores["Meia V"], valores["QR Code"],
                                         empresa))
        self._limpar_campos()

    def remover_selecionado(self):
        sel = self.tabela.selection()
        if not sel:
            messagebox.showinfo("Aviso", "Selecione uma linha para remover.")
            return
        if messagebox.askyesno("Confirmação", "Remover a linha selecionada?"):
            for item in sel:
                self.tabela.delete(item)

    def _obter_ou_criar_diretorio_saida(self):
        pasta_data = self.entry_data.get().strip() or datetime.now().strftime("%d/%m/%Y")
        pasta_data_fs = pasta_data.replace("/", "-")
        desktop = os.path.join(os.path.expanduser("~"), "Desktop")
        dir_saida = os.path.join(desktop, pasta_data_fs)
        os.makedirs(dir_saida, exist_ok=True)
        return dir_saida

    def _processar_registros_tabela(self):
        registros = []
        for it in self.tabela.get_children():
            v = list(self.tabela.item(it, "values"))
            try:
                nums = []
                for i in range(4, 9):
                    val_str = str(v[i]).split(' ')[0]
                    nums.append(int(val_str) if val_str and val_str != '-' else 0)
            except ValueError:
                messagebox.showerror("Erro de Dados", "Dados numéricos inválidos encontrados na tabela. Verifique as entradas.")
                return None
            if sum(nums) == 0:
                continue
            v[4:9] = nums
            registros.append(v)
        return registros

    def _calcular_totais(self, registros):
        chave_map = {}
        for r in registros:
            data, linha, prefixo, turno = r[0], r[1], r[2], r[3]
            vt, exp, meia, meiav, qr, empresa = r[4], r[5], r[6], r[7], r[8], r[9]
            key = (data, linha, prefixo, turno, empresa)
            if key not in chave_map:
                chave_map[key] = {t: 0 for t in TIPOS}
            chave_map[key]["VT"] += vt
            chave_map[key]["Exp"] += exp
            chave_map[key]["Meia"] += meia
            chave_map[key]["Meia V"] += meiav
            chave_map[key]["QR Code"] += qr

        total_geral = 0.0
        totais_tipo_qtd = {t: 0 for t in TIPOS}
        totais_tipo_val = {t: 0.0 for t in TIPOS}
        dados_tabela_formatados = []

        for (data, linha, prefixo, turno, empresa), somas in sorted(chave_map.items(), key=lambda x: x[0][2]):
            linha_vals = [data, linha, prefixo, turno]
            subtotal_rs = 0.0
            for t in TIPOS:
                qtd = somas[t]
                val = qtd * tarifa(linha, t)
                linha_vals.append(f"{qtd} ({brl(val)})" if qtd else "-")
                totais_tipo_qtd[t] += qtd
                totais_tipo_val[t] += val
                subtotal_rs += val
            linha_vals.extend([brl(subtotal_rs), empresa])
            total_geral += subtotal_rs
            dados_tabela_formatados.append(linha_vals)
        
        return dados_tabela_formatados, total_geral, totais_tipo_qtd, totais_tipo_val

    def exportar_pdf(self):
        itens = self.tabela.get_children()
        if not itens:
            messagebox.showwarning("Aviso", "Nenhum dado para exportar.")
            return

        registros = self._processar_registros_tabela()
        if registros is None: # Error occurred during processing
            return

        if not registros:
            messagebox.showwarning("Aviso", "Todos os lançamentos estão zerados ou contêm dados inválidos.")
            return

        dados_tabela_formatados, total_geral, totais_tipo_qtd, totais_tipo_val = self._calcular_totais(registros)

        dir_saida = self._obter_ou_criar_diretorio_saida()
        current_time_str = datetime.now().strftime("%H-%M-%S")
        nome_arquivo = os.path.join(dir_saida, f"Relatorio_{current_time_str}.pdf")
        estilos = getSampleStyleSheet()
        elementos = []
        doc = SimpleDocTemplate(nome_arquivo, pagesize=landscape(A4),
                                leftMargin=18, rightMargin=18, topMargin=18, bottomMargin=18)

        data_relatorio = self.entry_data.get().strip() or datetime.now().strftime("%d/%m/%Y")
        elementos.append(Paragraph(f"Relatório de Bilhetagem — {data_relatorio}", estilos["Title"]))
        elementos.append(Paragraph(f"Empresa: {self.combo_empresa.get()}", estilos["Heading2"]))
        elementos.append(Paragraph(f"Emitido em: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}", estilos["Normal"]))
        elementos.append(Spacer(1, 10))

        cab = ["Data", "Linha", "Prefixo", "Turno"] + TIPOS + ["Subtotal (R$)", "Empresa"]
        dados_tabela = [cab] + dados_tabela_formatados

        t1 = Table(dados_tabela, repeatRows=1, hAlign='CENTER')
        t1_style = TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
        ])
        for i, row in enumerate(dados_tabela[1:], start=1):
            turno = row[3]
            if turno == "1º Turno":
                t1_style.add("BACKGROUND", (0, i), (-1, i), colors.white)
            elif turno == "2º Turno":
                t1_style.add("BACKGROUND", (0, i), (-1, i), colors.lightgrey)
            elif turno == "3º Turno":
                t1_style.add("BACKGROUND", (0, i), (-1, i), colors.lavender)
            t1_style.add("ALIGN", (0, i), (3, i), "CENTER")
            t1_style.add("ALIGN", (4, i), (8, i), "RIGHT")
            t1_style.add("ALIGN", (9, i), (9, i), "RIGHT")
            t1_style.add("ALIGN", (10, i), (10, i), "CENTER")

        t1.setStyle(t1_style)
        elementos.append(Paragraph("Lançamentos Detalhados", estilos["Heading2"]))
        elementos.append(t1)
        elementos.append(Spacer(1, 12))

        elementos.append(Paragraph("Totais Gerais por Tipo", estilos["Heading2"]))
        tbl_totais = [["Tipo", "Quantidade", "Total (R$)"]]
        for t in TIPOS:
            tbl_totais.append([t, totais_tipo_qtd[t], brl(totais_tipo_val[t])])
        tbl_totais.append(["TOTAL GERAL", "", brl(total_geral)])

        t2 = Table(tbl_totais, repeatRows=1, hAlign='CENTER')
        t2.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("ALIGN", (2, 1), (2, -1), "RIGHT"),
            ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ]))
        elementos.append(t2)
        elementos.append(Spacer(1, 12))

        try:
            doc.build(elementos)
            messagebox.showinfo("Sucesso", f"PDF exportado para:\n{nome_arquivo}")
            webbrowser.open(f"file://{nome_arquivo}")
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao exportar PDF: {str(e)}")

    def exportar_excel(self):
        itens = self.tabela.get_children()
        if not itens:
            messagebox.showwarning("Aviso", "Nenhum dado para exportar.")
            return

        registros = self._processar_registros_tabela()
        if registros is None: # Error occurred during processing
            return

        if not registros:
            messagebox.showwarning("Aviso", "Todos os lançamentos estão zerados ou contêm dados inválidos.")
            return

        dados_tabela_formatados, total_geral, totais_tipo_qtd, totais_tipo_val = self._calcular_totais(registros)

        wb = Workbook()
        ws = wb.active
        ws.title = "Relatório de Bilhetagem"

        cab = ["Data", "Linha", "Prefixo", "Turno"] + TIPOS + ["Subtotal (R$)", "Empresa"]
        ws.append(cab)

        for col_idx, header in enumerate(cab, 1):
            ws.cell(row=1, column=col_idx).font = Font(bold=True)
            ws.cell(row=1, column=col_idx).fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            ws.cell(row=1, column=col_idx).alignment = Alignment(horizontal="center")

        for linha_vals in dados_tabela_formatados:
            ws.append(linha_vals)

        ultima_linha = ws.max_row + 2
        ws.cell(row=ultima_linha, column=1, value="Totais Gerais").font = Font(bold=True)
        for i, t in enumerate(TIPOS, start=5):
            ws.cell(row=ultima_linha, column=i, value=f"{totais_tipo_qtd[t]} ({brl(totais_tipo_val[t])})").font = Font(bold=True)

        ws.cell(row=ultima_linha+1, column=1, value="TOTAL GERAL").font = Font(bold=True)
        ws.cell(row=ultima_linha+1, column=6, value=brl(total_geral)).font = Font(bold=True)

        for col in ws.columns:
            max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_len + 2

        dir_saida = self._obter_ou_criar_diretorio_saida()
        current_time_str = datetime.now().strftime("%H-%M-%S")
        nome_arquivo = os.path.join(dir_saida, f"Relatorio_{current_time_str}.xlsx")
        try:
            wb.save(nome_arquivo)
            messagebox.showinfo("Sucesso", f"Excel exportado para:\n{nome_arquivo}")
            webbrowser.open(f"file://{nome_arquivo}")
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao exportar Excel: {str(e)}")


if __name__ == "__main__":
    root = tb.Window(themename="cosmo")
    app = ControleBilhetagemApp(root)
    root.mainloop()


